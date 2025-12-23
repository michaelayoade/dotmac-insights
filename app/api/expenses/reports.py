"""Expense reports and exports module."""
from __future__ import annotations

import io
from datetime import datetime, date
from decimal import Decimal
from typing import Any, Dict, List, Optional, cast

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, extract, case
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.auth import Require, Principal, get_current_principal
from app.models.expense_management import (
    ExpenseClaim,
    ExpenseClaimLine,
    ExpenseClaimStatus,
    CashAdvance,
    CashAdvanceStatus,
    ExpenseCategory,
    CorporateCard,
    CorporateCardTransaction,
    CardTransactionStatus,
)
from app.models.employee import Employee

# Optional Excel support
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Optional PDF support
try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False

router = APIRouter()


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def _format_date(d: Optional[Any]) -> str:
    """Format date for display."""
    if d is None:
        return ""
    if isinstance(d, datetime):
        return d.strftime("%Y-%m-%d")
    if isinstance(d, date):
        return d.strftime("%Y-%m-%d")
    return str(d)


def _format_number(value: Any) -> str:
    """Format number for display."""
    if value is None:
        return "0.00"
    if isinstance(value, (int, float, Decimal)):
        return f"{float(value):,.2f}"
    return str(value)


def _export_headers(base_filename: str, extension: str) -> Dict[str, str]:
    """Build Content-Disposition headers."""
    return {"Content-Disposition": f'attachment; filename="{base_filename}.{extension}"'}


# =============================================================================
# EXPORT STATUS
# =============================================================================

@router.get("/reports/status")
def get_expense_export_status() -> Dict[str, Any]:
    """Get export service availability."""
    return {
        "as_of": datetime.utcnow().isoformat() + "Z",
        "formats": {
            "csv": {"available": True},
            "excel": {"available": OPENPYXL_AVAILABLE, "requires": "openpyxl"},
            "pdf": {"available": WEASYPRINT_AVAILABLE, "requires": "weasyprint"},
        },
    }


# =============================================================================
# EXPENSE CLAIMS REPORT
# =============================================================================

@router.get("/reports/claims", dependencies=[Depends(Require("expenses:read"))])
def export_claims_report(
    format: str = Query("csv", description="Export format: csv, excel, or pdf"),
    start_date: Optional[str] = Query(None, description="Filter from date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="Filter to date (YYYY-MM-DD)"),
    status: Optional[str] = Query(None, description="Filter by status"),
    employee_id: Optional[int] = Query(None, description="Filter by employee"),
    include_lines: bool = Query(False, description="Include line item details"),
    filename: Optional[str] = Query(None, description="Custom filename"),
    db: Session = Depends(get_db),
    principal: Principal = Depends(get_current_principal),
):
    """Export expense claims report."""
    if format not in ("csv", "excel", "pdf"):
        raise HTTPException(status_code=400, detail="Format must be csv, excel, or pdf")

    if format == "excel" and not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=400, detail="Excel export not available (install openpyxl)")
    if format == "pdf" and not WEASYPRINT_AVAILABLE:
        raise HTTPException(status_code=400, detail="PDF export not available (install weasyprint)")

    # Build query
    query = db.query(ExpenseClaim).options(
        joinedload(ExpenseClaim.lines).joinedload(ExpenseClaimLine.category)
    ).order_by(ExpenseClaim.claim_date.desc())

    if start_date:
        query = query.filter(ExpenseClaim.claim_date >= start_date)
    if end_date:
        query = query.filter(ExpenseClaim.claim_date <= end_date)
    if status:
        try:
            query = query.filter(ExpenseClaim.status == ExpenseClaimStatus(status))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid status filter")
    if employee_id:
        query = query.filter(ExpenseClaim.employee_id == employee_id)

    claims = query.all()

    # Build report data
    data = _build_claims_report_data(claims, include_lines, start_date, end_date)
    base_filename = filename or f"expense_claims_{datetime.now().strftime('%Y%m%d')}"

    if format == "csv":
        csv_content = _claims_to_csv(data, include_lines)
        return StreamingResponse(
            iter([csv_content]),
            media_type="text/csv",
            headers=_export_headers(base_filename, "csv"),
        )
    elif format == "excel":
        excel_content = _claims_to_excel(data, include_lines)
        return StreamingResponse(
            iter([excel_content]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=_export_headers(base_filename, "xlsx"),
        )
    else:  # pdf
        pdf_content = _claims_to_pdf(data, include_lines)
        return StreamingResponse(
            iter([pdf_content]),
            media_type="application/pdf",
            headers=_export_headers(base_filename, "pdf"),
        )


def _build_claims_report_data(claims: List[ExpenseClaim], include_lines: bool, start_date: Optional[str], end_date: Optional[str]) -> Dict[str, Any]:
    """Build structured report data from claims."""
    rows = []
    total_claimed = Decimal("0")
    total_approved = Decimal("0")

    status_counts: Dict[str, int] = {}

    for claim in claims:
        status_key = claim.status.value if claim.status else "unknown"
        status_counts[status_key] = status_counts.get(status_key, 0) + 1
        total_claimed += claim.total_claimed_amount or Decimal("0")
        total_approved += claim.total_sanctioned_amount or Decimal("0")

        row = {
            "claim_number": claim.claim_number or str(claim.id),
            "title": claim.title,
            "employee_id": claim.employee_id,
            "claim_date": _format_date(claim.claim_date),
            "status": status_key,
            "currency": claim.currency,
            "total_claimed": float(claim.total_claimed_amount or 0),
            "approved_amount": float(claim.total_sanctioned_amount or 0),
            "line_count": len(claim.lines) if claim.lines else 0,
            "submitted_at": _format_date(claim.submitted_at),
            "approved_at": _format_date(claim.approved_at),
        }

        if include_lines and claim.lines:
            row["lines"] = [
                {
                    "category": line.category.name if line.category else "Uncategorized",
                    "description": line.description or "",
                    "expense_date": _format_date(line.expense_date),
                    "amount": float(line.claimed_amount or 0),
                    "currency": line.currency,
                    "receipt_attached": bool(line.has_receipt),
                }
                for line in claim.lines
            ]

        rows.append(row)

    return {
        "report_title": "Expense Claims Report",
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "period": {
            "start": start_date or "Beginning",
            "end": end_date or "Present",
        },
        "summary": {
            "total_claims": len(claims),
            "total_claimed": float(total_claimed),
            "total_approved": float(total_approved),
            "by_status": status_counts,
        },
        "claims": rows,
    }


def _claims_to_csv(data: Dict[str, Any], include_lines: bool) -> str:
    """Convert claims report to CSV using Jinja2 template."""
    from app.templates.environment import get_template_env

    env = get_template_env()
    template = env.get_template("reports/expenses/claims.csv.j2")
    return template.render(
        generated_at=data["generated_at"],
        period=data["period"],
        summary=data["summary"],
        claims=data["claims"],
        include_lines=include_lines,
    )


def _claims_to_excel(data: Dict[str, Any], include_lines: bool) -> bytes:
    """Convert claims report to Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Claims Report"

    # Styles
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    table_header_font = Font(bold=True)
    table_header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    currency_font = Font(name="Consolas")

    row = 1

    # Title
    ws.cell(row=row, column=1, value="Expense Claims Report").font = header_font
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {data['generated_at']}")
    row += 1
    ws.cell(row=row, column=1, value=f"Period: {data['period']['start']} to {data['period']['end']}")
    row += 2

    # Summary section
    ws.cell(row=row, column=1, value="Summary").font = subheader_font
    row += 1
    summary = data["summary"]
    ws.cell(row=row, column=1, value="Total Claims")
    ws.cell(row=row, column=2, value=summary["total_claims"])
    row += 1
    ws.cell(row=row, column=1, value="Total Claimed")
    ws.cell(row=row, column=2, value=summary["total_claimed"]).font = currency_font
    row += 1
    ws.cell(row=row, column=1, value="Total Approved")
    ws.cell(row=row, column=2, value=summary["total_approved"]).font = currency_font
    row += 2

    # Data table
    ws.cell(row=row, column=1, value="Claims Detail").font = subheader_font
    row += 1

    if include_lines:
        headers = ["Claim #", "Title", "Employee ID", "Date", "Status", "Category", "Description", "Expense Date", "Amount", "Currency"]
    else:
        headers = ["Claim #", "Title", "Employee ID", "Date", "Status", "Lines", "Total Claimed", "Approved", "Currency"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = table_header_font
        cell.fill = table_header_fill
    row += 1

    if include_lines:
        for claim in data["claims"]:
            lines = claim.get("lines", [])
            if lines:
                for i, line in enumerate(lines):
                    if i == 0:
                        ws.cell(row=row, column=1, value=claim["claim_number"])
                        ws.cell(row=row, column=2, value=claim["title"])
                        ws.cell(row=row, column=3, value=claim["employee_id"])
                        ws.cell(row=row, column=4, value=claim["claim_date"])
                        ws.cell(row=row, column=5, value=claim["status"])
                    ws.cell(row=row, column=6, value=line["category"])
                    ws.cell(row=row, column=7, value=line["description"])
                    ws.cell(row=row, column=8, value=line["expense_date"])
                    ws.cell(row=row, column=9, value=line["amount"]).font = currency_font
                    ws.cell(row=row, column=10, value=line["currency"])
                    row += 1
            else:
                ws.cell(row=row, column=1, value=claim["claim_number"])
                ws.cell(row=row, column=2, value=claim["title"])
                ws.cell(row=row, column=3, value=claim["employee_id"])
                ws.cell(row=row, column=4, value=claim["claim_date"])
                ws.cell(row=row, column=5, value=claim["status"])
                ws.cell(row=row, column=9, value=claim["total_claimed"]).font = currency_font
                ws.cell(row=row, column=10, value=claim["currency"])
                row += 1
    else:
        for claim in data["claims"]:
            ws.cell(row=row, column=1, value=claim["claim_number"])
            ws.cell(row=row, column=2, value=claim["title"])
            ws.cell(row=row, column=3, value=claim["employee_id"])
            ws.cell(row=row, column=4, value=claim["claim_date"])
            ws.cell(row=row, column=5, value=claim["status"])
            ws.cell(row=row, column=6, value=claim["line_count"])
            ws.cell(row=row, column=7, value=claim["total_claimed"]).font = currency_font
            ws.cell(row=row, column=8, value=claim["approved_amount"]).font = currency_font
            ws.cell(row=row, column=9, value=claim["currency"])
            row += 1

    # Auto-size columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions["B"].width = 30  # Title
    ws.column_dimensions["G"].width = 30 if include_lines else 15  # Description

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _claims_to_pdf(data: Dict[str, Any], include_lines: bool) -> bytes:
    """Convert claims report to PDF using Jinja2 template."""
    from app.templates.environment import get_template_env
    from pathlib import Path

    env = get_template_env()
    template = env.get_template("reports/expenses/claims.html.j2")

    html_content = template.render(
        period=data["period"],
        summary=data["summary"],
        claims=data["claims"],
        include_lines=include_lines,
        generated_at=data["generated_at"],
    )

    # Load CSS from template file
    css_path = Path(__file__).parent.parent.parent / "templates" / "reports" / "expenses" / "_styles.css"
    css = css_path.read_text()

    html = HTML(string=html_content)
    return cast(bytes, html.write_pdf(stylesheets=[CSS(string=css)]))


# =============================================================================
# CASH ADVANCES REPORT
# =============================================================================

@router.get("/reports/advances", dependencies=[Depends(Require("expenses:read"))])
def export_advances_report(
    format: str = Query("csv", description="Export format: csv, excel, or pdf"),
    start_date: Optional[str] = Query(None, description="Filter from date"),
    end_date: Optional[str] = Query(None, description="Filter to date"),
    status: Optional[str] = Query(None, description="Filter by status"),
    employee_id: Optional[int] = Query(None, description="Filter by employee"),
    filename: Optional[str] = Query(None, description="Custom filename"),
    db: Session = Depends(get_db),
    principal: Principal = Depends(get_current_principal),
):
    """Export cash advances report."""
    if format not in ("csv", "excel", "pdf"):
        raise HTTPException(status_code=400, detail="Format must be csv, excel, or pdf")

    if format == "excel" and not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=400, detail="Excel export not available")
    if format == "pdf" and not WEASYPRINT_AVAILABLE:
        raise HTTPException(status_code=400, detail="PDF export not available")

    # Build query
    query = db.query(CashAdvance).order_by(CashAdvance.request_date.desc())

    if start_date:
        query = query.filter(CashAdvance.request_date >= start_date)
    if end_date:
        query = query.filter(CashAdvance.request_date <= end_date)
    if status:
        try:
            query = query.filter(CashAdvance.status == CashAdvanceStatus(status))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid status filter")
    if employee_id:
        query = query.filter(CashAdvance.employee_id == employee_id)

    advances = query.all()
    data = _build_advances_report_data(advances, start_date, end_date)
    base_filename = filename or f"cash_advances_{datetime.now().strftime('%Y%m%d')}"

    if format == "csv":
        csv_content = _advances_to_csv(data)
        return StreamingResponse(
            iter([csv_content]),
            media_type="text/csv",
            headers=_export_headers(base_filename, "csv"),
        )
    elif format == "excel":
        excel_content = _advances_to_excel(data)
        return StreamingResponse(
            iter([excel_content]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=_export_headers(base_filename, "xlsx"),
        )
    else:  # pdf
        pdf_content = _advances_to_pdf(data)
        return StreamingResponse(
            iter([pdf_content]),
            media_type="application/pdf",
            headers=_export_headers(base_filename, "pdf"),
        )


def _build_advances_report_data(advances: List[CashAdvance], start_date: Optional[str], end_date: Optional[str]) -> Dict[str, Any]:
    """Build structured report data from advances."""
    rows = []
    total_requested = Decimal("0")
    total_disbursed = Decimal("0")
    total_settled = Decimal("0")
    total_outstanding = Decimal("0")

    status_counts: Dict[str, int] = {}

    for adv in advances:
        status_key = adv.status.value if adv.status else "unknown"
        status_counts[status_key] = status_counts.get(status_key, 0) + 1
        total_requested += adv.requested_amount or Decimal("0")
        total_disbursed += adv.disbursed_amount or Decimal("0")
        total_settled += adv.settled_amount or Decimal("0")
        total_outstanding += adv.outstanding_amount or Decimal("0")

        rows.append({
            "advance_number": adv.advance_number or str(adv.id),
            "employee_id": adv.employee_id,
            "purpose": adv.purpose,
            "request_date": _format_date(adv.request_date),
            "required_date": _format_date(adv.required_by_date),
            "status": status_key,
            "currency": adv.currency,
            "requested_amount": float(adv.requested_amount or 0),
            "disbursed_amount": float(adv.disbursed_amount or 0),
            "settled_amount": float(adv.settled_amount or 0),
            "outstanding_balance": float(adv.outstanding_amount or 0),
            "disbursed_at": _format_date(adv.disbursed_at),
        })

    return {
        "report_title": "Cash Advances Report",
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "period": {
            "start": start_date or "Beginning",
            "end": end_date or "Present",
        },
        "summary": {
            "total_advances": len(advances),
            "total_requested": float(total_requested),
            "total_disbursed": float(total_disbursed),
            "total_settled": float(total_settled),
            "total_outstanding": float(total_outstanding),
            "by_status": status_counts,
        },
        "advances": rows,
    }


def _advances_to_csv(data: Dict[str, Any]) -> str:
    """Convert advances report to CSV using Jinja2 template."""
    from app.templates.environment import get_template_env

    env = get_template_env()
    template = env.get_template("reports/expenses/advances.csv.j2")
    return template.render(
        generated_at=data["generated_at"],
        period=data["period"],
        summary=data["summary"],
        advances=data["advances"],
    )


def _advances_to_excel(data: Dict[str, Any]) -> bytes:
    """Convert advances report to Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Advances Report"

    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    table_header_font = Font(bold=True)
    table_header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    currency_font = Font(name="Consolas")

    row = 1
    ws.cell(row=row, column=1, value="Cash Advances Report").font = header_font
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {data['generated_at']}")
    row += 1
    ws.cell(row=row, column=1, value=f"Period: {data['period']['start']} to {data['period']['end']}")
    row += 2

    summary = data["summary"]
    ws.cell(row=row, column=1, value="Summary").font = subheader_font
    row += 1
    for label, value in [
        ("Total Advances", summary["total_advances"]),
        ("Total Requested", summary["total_requested"]),
        ("Total Disbursed", summary["total_disbursed"]),
        ("Total Settled", summary["total_settled"]),
        ("Total Outstanding", summary["total_outstanding"]),
    ]:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        if isinstance(value, float):
            cell.font = currency_font
        row += 1
    row += 1

    headers = ["Advance #", "Employee ID", "Purpose", "Request Date", "Status", "Currency", "Requested", "Disbursed", "Settled", "Outstanding"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = table_header_font
        cell.fill = table_header_fill
    row += 1

    for adv in data["advances"]:
        ws.cell(row=row, column=1, value=adv["advance_number"])
        ws.cell(row=row, column=2, value=adv["employee_id"])
        ws.cell(row=row, column=3, value=adv["purpose"])
        ws.cell(row=row, column=4, value=adv["request_date"])
        ws.cell(row=row, column=5, value=adv["status"])
        ws.cell(row=row, column=6, value=adv["currency"])
        ws.cell(row=row, column=7, value=adv["requested_amount"]).font = currency_font
        ws.cell(row=row, column=8, value=adv["disbursed_amount"]).font = currency_font
        ws.cell(row=row, column=9, value=adv["settled_amount"]).font = currency_font
        ws.cell(row=row, column=10, value=adv["outstanding_balance"]).font = currency_font
        row += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["C"].width = 30

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _advances_to_pdf(data: Dict[str, Any]) -> bytes:
    """Convert advances report to PDF using Jinja2 template."""
    from app.templates.environment import get_template_env
    from pathlib import Path

    env = get_template_env()
    template = env.get_template("reports/expenses/advances.html.j2")

    html_content = template.render(
        period=data["period"],
        summary=data["summary"],
        advances=data["advances"],
        generated_at=data["generated_at"],
    )

    # Load CSS from template file
    css_path = Path(__file__).parent.parent.parent / "templates" / "reports" / "expenses" / "_styles.css"
    css = css_path.read_text()

    html = HTML(string=html_content)
    return cast(bytes, html.write_pdf(stylesheets=[CSS(string=css)]))


# =============================================================================
# EXPENSE SUMMARY REPORT (COMBINED)
# =============================================================================

@router.get("/reports/summary", dependencies=[Depends(Require("expenses:read"))])
def get_expense_summary_report(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Get expense summary statistics (JSON)."""
    # Claims summary
    claims_query = db.query(ExpenseClaim)
    advances_query = db.query(CashAdvance)

    if start_date:
        claims_query = claims_query.filter(ExpenseClaim.claim_date >= start_date)
        advances_query = advances_query.filter(CashAdvance.request_date >= start_date)
    if end_date:
        claims_query = claims_query.filter(ExpenseClaim.claim_date <= end_date)
        advances_query = advances_query.filter(CashAdvance.request_date <= end_date)

    # Aggregate claims
    claims_agg_query = db.query(
        func.count(ExpenseClaim.id).label("count"),
        func.sum(ExpenseClaim.total_claimed_amount).label("total_claimed"),
        func.sum(ExpenseClaim.total_sanctioned_amount).label("total_approved"),
    )
    if start_date:
        claims_agg_query = claims_agg_query.filter(ExpenseClaim.claim_date >= start_date)
    if end_date:
        claims_agg_query = claims_agg_query.filter(ExpenseClaim.claim_date <= end_date)
    claims_agg = claims_agg_query.first()

    # Aggregate advances
    advances_agg_query = db.query(
        func.count(CashAdvance.id).label("count"),
        func.sum(CashAdvance.requested_amount).label("total_requested"),
        func.sum(CashAdvance.disbursed_amount).label("total_disbursed"),
        func.sum(CashAdvance.outstanding_amount).label("total_outstanding"),
    )
    if start_date:
        advances_agg_query = advances_agg_query.filter(CashAdvance.request_date >= start_date)
    if end_date:
        advances_agg_query = advances_agg_query.filter(CashAdvance.request_date <= end_date)
    advances_agg = advances_agg_query.first()

    # Status breakdown for claims
    claims_by_status_query = db.query(
        ExpenseClaim.status,
        func.count(ExpenseClaim.id).label("count"),
        func.sum(ExpenseClaim.total_claimed_amount).label("amount"),
    )
    if start_date:
        claims_by_status_query = claims_by_status_query.filter(ExpenseClaim.claim_date >= start_date)
    if end_date:
        claims_by_status_query = claims_by_status_query.filter(ExpenseClaim.claim_date <= end_date)
    claims_by_status = claims_by_status_query.group_by(ExpenseClaim.status).all()

    # Top expense categories
    top_categories_query = db.query(
        ExpenseCategory.name,
        func.count(ExpenseClaimLine.id).label("count"),
        func.sum(ExpenseClaimLine.claimed_amount).label("total"),
    ).join(
        ExpenseClaimLine, ExpenseClaimLine.category_id == ExpenseCategory.id
    ).join(
        ExpenseClaim, ExpenseClaimLine.expense_claim_id == ExpenseClaim.id
    )
    if start_date:
        top_categories_query = top_categories_query.filter(ExpenseClaim.claim_date >= start_date)
    if end_date:
        top_categories_query = top_categories_query.filter(ExpenseClaim.claim_date <= end_date)
    top_categories = top_categories_query.group_by(
        ExpenseCategory.name
    ).order_by(
        func.sum(ExpenseClaimLine.claimed_amount).desc()
    ).limit(10).all()

    return {
        "period": {
            "start": start_date or "All time",
            "end": end_date or "Present",
        },
        "claims": {
            "count": claims_agg.count if claims_agg else 0,
            "total_claimed": float(claims_agg.total_claimed or 0) if claims_agg else 0,
            "total_approved": float(claims_agg.total_approved or 0) if claims_agg else 0,
            "by_status": [
                {"status": row.status.value if row.status else "unknown", "count": row.count, "amount": float(row.amount or 0)}
                for row in claims_by_status
            ],
        },
        "advances": {
            "count": advances_agg.count if advances_agg else 0,
            "total_requested": float(advances_agg.total_requested or 0) if advances_agg else 0,
            "total_disbursed": float(advances_agg.total_disbursed or 0) if advances_agg else 0,
            "total_outstanding": float(advances_agg.total_outstanding or 0) if advances_agg else 0,
        },
        "top_categories": [
            {"category": row.name, "count": row.count, "total": float(row.total or 0)}
            for row in top_categories
        ],
        "generated_at": datetime.utcnow().isoformat() + "Z",
    }


# =============================================================================
# CARD TRANSACTIONS REPORT
# =============================================================================

@router.get("/reports/transactions", dependencies=[Depends(Require("expenses:read"))])
def export_transactions_report(
    format: str = Query("csv", description="Export format: csv or excel"),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    card_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    filename: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    principal: Principal = Depends(get_current_principal),
):
    """Export corporate card transactions report."""
    if format not in ("csv", "excel"):
        raise HTTPException(status_code=400, detail="Format must be csv or excel")
    if format == "excel" and not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=400, detail="Excel export not available")

    query = db.query(CorporateCardTransaction).join(
        CorporateCard, CorporateCardTransaction.card_id == CorporateCard.id
    ).order_by(CorporateCardTransaction.transaction_date.desc())

    if start_date:
        query = query.filter(CorporateCardTransaction.transaction_date >= start_date)
    if end_date:
        query = query.filter(CorporateCardTransaction.transaction_date <= end_date)
    if card_id:
        query = query.filter(CorporateCardTransaction.card_id == card_id)
    if status:
        try:
            query = query.filter(CorporateCardTransaction.status == CardTransactionStatus(status))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid status")

    transaction_rows = query.limit(5000).all()  # Limit for performance

    # Get card names
    card_map = {c.id: c.card_name for c in db.query(CorporateCard).all()}

    data: Dict[str, Any] = {
        "report_title": "Card Transactions Report",
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "period": {"start": start_date or "All time", "end": end_date or "Present"},
        "transactions": [
            {
                "card_name": card_map.get(t.card_id, f"Card {t.card_id}"),
                "transaction_date": _format_date(t.transaction_date),
                "merchant": t.merchant_name or "",
                "description": t.description or "",
                "amount": float(t.amount or 0),
                "currency": t.currency,
                "status": t.status.value if t.status else "unknown",
                "reference": t.transaction_reference or "",
            }
            for t in transaction_rows
        ],
    }

    base_filename = filename or f"card_transactions_{datetime.now().strftime('%Y%m%d')}"

    transaction_items = cast(List[Dict[str, Any]], data["transactions"])

    if format == "csv":
        from app.templates.environment import get_template_env

        env = get_template_env()
        template = env.get_template("reports/expenses/transactions.csv.j2")
        csv_content = template.render(
            generated_at=data["generated_at"],
            transactions=transaction_items,
        )
        return StreamingResponse(
            iter([csv_content]),
            media_type="text/csv",
            headers=_export_headers(base_filename, "csv"),
        )
    else:  # excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        ws.cell(row=1, column=1, value="Card Transactions Report").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Generated: {data['generated_at']}")

        headers = ["Card", "Date", "Merchant", "Description", "Amount", "Currency", "Status", "Reference"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=h).font = Font(bold=True)

        for i, t in enumerate(transaction_items, 5):
            ws.cell(row=i, column=1, value=t["card_name"])
            ws.cell(row=i, column=2, value=t["transaction_date"])
            ws.cell(row=i, column=3, value=t["merchant"])
            ws.cell(row=i, column=4, value=t["description"])
            ws.cell(row=i, column=5, value=t["amount"]).font = Font(name="Consolas")
            ws.cell(row=i, column=6, value=t["currency"])
            ws.cell(row=i, column=7, value=t["status"])
            ws.cell(row=i, column=8, value=t["reference"])

        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 30

        output_bytes = io.BytesIO()
        wb.save(output_bytes)
        return StreamingResponse(
            iter([output_bytes.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=_export_headers(base_filename, "xlsx"),
        )


